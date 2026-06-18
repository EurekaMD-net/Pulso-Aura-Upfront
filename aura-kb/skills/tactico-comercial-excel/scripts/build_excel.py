#!/usr/bin/env python3
"""
TÁCTICO COMERCIAL EXCEL — Build Script
Genera el archivo .xlsx de propuesta comercial de medios con formato profesional.

USO:
    python build_excel.py --data propuesta.json --output propuesta_comercial.xlsx

FORMATO DE DATOS (propuesta.json):
{
  "cliente": "Nombre del Cliente",
  "marca": "Nombre de la Marca",
  "periodo": "Q2 2025 / Mayo-Junio 2025",
  "presupuesto_total": 5000000,
  "objetivo": "Lanzamiento de producto, incrementar awareness y consideración",
  "elaborado_por": "Área Comercial / KAM",
  "fecha": "Mayo 2025",
  
  "resumen_ejecutivo": "Descripción de la propuesta en 2-3 oraciones.",
  
  "filas": [
    {
      "medio": "TV Lineal",
      "propiedad": "Azteca Uno",
      "formato": "Spot 20\"",
      "descripcion": "Spot en bloque comercial en horario prime time.",
      "objetivo": "Awareness masivo",
      "etapa_funnel": "Top",
      "audiencia": "A18-49, NSE ABC+, hogares con TV abierta",
      "entregable": "Spots al aire",
      "cantidad": 120,
      "unidad_compra": "Spots",
      "precio_unitario": 45000,
      "kpi_primario": "GRP / Alcance",
      "kpi_secundario": "Frecuencia media",
      "racional": "Azteca Uno es la señal de mayor alcance en TV abierta mexicana...",
      "supuestos": "Precios de lista. Sujeto a disponibilidad en prime time."
    }
  ],
  
  "supuestos_generales": [
    "Precios en MXN, incluyen IVA.",
    "Propuesta sujeta a disponibilidad de inventario.",
    "CPMs son estimados de referencia y pueden variar por negociación."
  ],
  
  "notas_comerciales": [
    "Propuesta vigente hasta el 30 de mayo de 2025.",
    "Mínimo de compra en CTV: 500,000 impresiones por campaña.",
    "Se requieren creatividades 5 días hábiles antes del inicio."
  ]
}
"""

import json
import sys
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─── PALETA DE COLORES ───────────────────────────────────────────────────────
DARK_BLUE    = "1A2744"   # Header principal
MID_BLUE     = "2C4A8C"   # Subheaders
ACCENT_BLUE  = "4472C4"   # Acentos
LIGHT_BLUE   = "D9E2F3"   # Fondos alternos filas
VERY_LIGHT   = "F2F5FB"   # Fondo alterno claro
ORANGE       = "E8834A"   # Totales / highlights
GOLD         = "F5A623"   # KPI highlight
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F0F0F0"
MID_GRAY     = "BFBFBF"
DARK_GRAY    = "595959"

# Colores por medio
MEDIO_COLORS = {
    "TV Lineal":   "FF6B6B",
    "CTV":         "4ECDC4",
    "Digital":     "45B7D1",
    "Radio":       "96CEB4",
    "OOH":         "DDA0DD",
    "Prensa":      "F0A500",
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def make_font(bold=False, size=10, color=None, name="Calibri"):
    return Font(bold=bold, size=size, color=color or "000000", name=name)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_align(horizontal="left", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def set_cell(ws, row, col, value, bold=False, size=10, bg=None, fg=None,
             align="left", wrap=True, num_format=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, color=fg or "000000", name="Calibri")
    if bg:
        cell.fill = make_fill(bg)
    cell.alignment = make_align(align, wrap=wrap)
    if border:
        cell.border = make_border()
    if num_format:
        cell.number_format = num_format
    return cell

def money_format(ws, row, col, value, bg=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '#,##0'
    cell.font = Font(bold=bold, size=10, name="Calibri")
    if bg:
        cell.fill = make_fill(bg)
    cell.alignment = make_align("right", wrap=False)
    cell.border = make_border()
    return cell


# ─── PESTAÑA 1: RESUMEN EJECUTIVO ────────────────────────────────────────────

def build_resumen(wb, data):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False

    # Configurar columnas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    row = 1

    # Header principal
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 45
    c = ws.cell(row=row, column=1, value="PROPUESTA COMERCIAL TÁCTICA DE MEDIOS")
    c.font = Font(bold=True, size=18, color=WHITE, name="Calibri")
    c.fill = make_fill(DARK_BLUE)
    c.alignment = make_align("center", wrap=False)
    row += 1

    # Datos del cliente
    meta_fields = [
        ("Cliente", data.get("cliente", "—")),
        ("Marca", data.get("marca", "—")),
        ("Período", data.get("periodo", "—")),
        ("Elaborado por", data.get("elaborado_por", "—")),
        ("Fecha", data.get("fecha", datetime.now().strftime("%B %Y"))),
    ]

    for label, value in meta_fields:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"B{row}:E{row}")
        set_cell(ws, row, 1, label, bold=True, bg=LIGHT_BLUE, size=10)
        set_cell(ws, row, 2, value, bg=WHITE, size=10)
        row += 1

    row += 1  # espacio

    # Objetivo
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 18
    set_cell(ws, row, 1, "OBJETIVO DE LA CAMPAÑA", bold=True, size=11,
             bg=MID_BLUE, fg=WHITE, align="center")
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 36
    set_cell(ws, row, 1, data.get("objetivo", "—"), bg=VERY_LIGHT, size=10, wrap=True)
    row += 1

    if data.get("resumen_ejecutivo"):
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 54
        set_cell(ws, row, 1, data["resumen_ejecutivo"], bg=WHITE, size=10, wrap=True)
        row += 1

    row += 1  # espacio

    # Tabla resumen por medio
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 18
    set_cell(ws, row, 1, "RESUMEN DE INVERSIÓN POR MEDIO", bold=True, size=11,
             bg=MID_BLUE, fg=WHITE, align="center")
    row += 1

    # Header tabla
    headers_resumen = ["Medio", "Formatos / Propiedades", "Etapa Funnel", "Inversión (MXN)", "KPI Principal"]
    widths = [30, 50, 20, 18, 20]
    for i, h in enumerate(headers_resumen):
        set_cell(ws, row, i+1, h, bold=True, bg=ACCENT_BLUE, fg=WHITE, size=10, align="center")
    row += 1

    # Agrupar por medio
    filas = data.get("filas", [])
    medios_data = {}
    for f in filas:
        medio = f.get("medio", "Otro")
        if medio not in medios_data:
            medios_data[medio] = {"inversion": 0, "formatos": set(), "etapas": set(), "kpis": set()}
        medios_data[medio]["inversion"] += f.get("cantidad", 0) * f.get("precio_unitario", 0)
        medios_data[medio]["formatos"].add(f.get("formato", ""))
        medios_data[medio]["etapas"].add(f.get("etapa_funnel", ""))
        medios_data[medio]["kpis"].add(f.get("kpi_primario", ""))

    total_inversion = 0
    alt = True
    for medio, md in medios_data.items():
        bg = LIGHT_BLUE if alt else VERY_LIGHT
        alt = not alt
        formatos_str = " / ".join(list(md["formatos"])[:3])
        etapas_str = " + ".join(sorted(md["etapas"]))
        kpis_str = " / ".join(list(md["kpis"])[:2])

        set_cell(ws, row, 1, medio, bold=True, bg=bg, size=10)
        set_cell(ws, row, 2, formatos_str, bg=bg, size=9)
        set_cell(ws, row, 3, etapas_str, bg=bg, size=9, align="center")
        money_format(ws, row, 4, md["inversion"], bg=bg, bold=True)
        set_cell(ws, row, 5, kpis_str, bg=bg, size=9)
        total_inversion += md["inversion"]
        row += 1

    # Total
    ws.row_dimensions[row].height = 22
    set_cell(ws, row, 1, "INVERSIÓN TOTAL", bold=True, size=11, bg=ORANGE, fg=WHITE)
    ws.merge_cells(f"B{row}:C{row}")
    set_cell(ws, row, 2, "", bg=ORANGE)
    money_format(ws, row, 4, total_inversion, bg=ORANGE, bold=True)
    ws.cell(row=row, column=4).font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    ws.cell(row=row, column=4).fill = make_fill(ORANGE)
    set_cell(ws, row, 5, "", bg=ORANGE)
    row += 2

    # Presupuesto del cliente si existe
    presupuesto = data.get("presupuesto_total", 0)
    if presupuesto:
        pct = (total_inversion / presupuesto * 100) if presupuesto else 0
        set_cell(ws, row, 1, f"Presupuesto disponible: ${presupuesto:,.0f} MXN   |   Utilización: {pct:.1f}%",
                 bold=False, size=9, bg=LIGHT_GRAY, wrap=False)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1


# ─── PESTAÑA 2: PROPUESTA TÁCTICA ────────────────────────────────────────────

def build_tactica(wb, data):
    ws = wb.create_sheet("Propuesta Táctica")
    ws.sheet_view.showGridLines = False

    # Anchos de columna
    col_widths = {
        "A": 12,   # Medio
        "B": 20,   # Propiedad
        "C": 20,   # Formato
        "D": 28,   # Descripción
        "E": 18,   # Objetivo / Rol
        "F": 10,   # Funnel
        "G": 22,   # Audiencia
        "H": 18,   # Entregable
        "I": 10,   # Cantidad
        "J": 14,   # Unidad
        "K": 14,   # Precio Unit
        "L": 16,   # Inversión Total
        "M": 14,   # CPM Estimado
        "N": 18,   # KPI Primario
        "O": 18,   # KPI Secundario
        "P": 30,   # Racional
        "Q": 25,   # Supuestos
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    row = 1

    # Título
    ws.merge_cells("A1:Q1")
    ws.row_dimensions[1].height = 35
    c = ws.cell(row=1, column=1, value="PLAN TÁCTICO COMERCIAL — DETALLE POR ACCIÓN")
    c.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill = make_fill(DARK_BLUE)
    c.alignment = make_align("center")
    row += 1

    # Info del cliente en 1 fila
    ws.merge_cells(f"A{row}:Q{row}")
    ws.row_dimensions[row].height = 18
    info_text = f"Cliente: {data.get('cliente','—')}  |  Marca: {data.get('marca','—')}  |  Período: {data.get('periodo','—')}"
    c = ws.cell(row=row, column=1, value=info_text)
    c.font = Font(bold=False, size=9, color=DARK_GRAY, name="Calibri")
    c.fill = make_fill(LIGHT_GRAY)
    c.alignment = make_align("left", wrap=False)
    row += 1

    # Encabezados
    headers = [
        "Medio", "Propiedad /\nPlataforma", "Formato", "Descripción",
        "Objetivo /\nRol", "Funnel", "Audiencia /\nContexto",
        "Entregable", "Cantidad", "Unidad\nCompra", "Precio\nUnitario",
        "Inversión\nTotal (MXN)", "CPM\nEstimado",
        "KPI Primario", "KPI Secundario", "Racional Comercial",
        "Supuestos /\nRestricciones"
    ]
    ws.row_dimensions[row].height = 36
    for i, h in enumerate(headers):
        set_cell(ws, row, i+1, h, bold=True, bg=DARK_BLUE, fg=WHITE,
                 size=9, align="center")
    row += 1
    ws.freeze_panes = ws.cell(row=row, column=1)  # Freeze header

    # Filas de datos
    filas = data.get("filas", [])
    alt = True
    last_medio = None
    total_inversion = 0

    for fila in filas:
        medio = fila.get("medio", "")
        es_nuevo_medio = (medio != last_medio)

        # Separador de medio (si cambia)
        if es_nuevo_medio and last_medio is not None:
            ws.row_dimensions[row].height = 6
            for col in range(1, 18):
                c = ws.cell(row=row, column=col)
                c.fill = make_fill(DARK_BLUE)
            row += 1
            alt = True

        bg = LIGHT_BLUE if alt else VERY_LIGHT
        alt = not alt
        ws.row_dimensions[row].height = 48

        inversion = fila.get("cantidad", 0) * fila.get("precio_unitario", 0)
        total_inversion += inversion

        # Color del medio
        medio_color = MEDIO_COLORS.get(medio, ACCENT_BLUE)

        # Columna MEDIO con color especial
        c = ws.cell(row=row, column=1, value=medio)
        c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.fill = make_fill(medio_color)
        c.alignment = make_align("center")
        c.border = make_border()

        valores = [
            fila.get("propiedad", ""),
            fila.get("formato", ""),
            fila.get("descripcion", ""),
            fila.get("objetivo", ""),
            fila.get("etapa_funnel", ""),
            fila.get("audiencia", ""),
            fila.get("entregable", ""),
            fila.get("cantidad", 0),
            fila.get("unidad_compra", ""),
            fila.get("precio_unitario", 0),
            inversion,
            fila.get("cpm_estimado", "—"),
            fila.get("kpi_primario", ""),
            fila.get("kpi_secundario", ""),
            fila.get("racional", ""),
            fila.get("supuestos", ""),
        ]

        for i, val in enumerate(valores):
            col = i + 2
            if col in (11, 12):  # precio y inversión — formato dinero
                money_format(ws, row, col, val, bg=bg, bold=(col == 12))
            elif col == 7:  # etapa funnel — centrado con color
                funnel_bg = {"Top": "FFE4B5", "Mid": "B8E1F5", "Bottom": "C8F5C8"}.get(str(val), bg)
                set_cell(ws, row, col, val, bg=funnel_bg, size=9, align="center")
            else:
                set_cell(ws, row, col, val, bg=bg, size=9)

        last_medio = medio
        row += 1

    # Fila de TOTAL
    ws.row_dimensions[row].height = 24
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row=row, column=1, value="INVERSIÓN TOTAL DE LA PROPUESTA")
    c.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    c.fill = make_fill(ORANGE)
    c.alignment = make_align("right", wrap=False)
    c.border = make_border()
    money_format(ws, row, 12, total_inversion, bg=ORANGE, bold=True)
    ws.cell(row=row, column=12).font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    ws.cell(row=row, column=12).fill = make_fill(ORANGE)
    for col in range(13, 18):
        set_cell(ws, row, col, "", bg=ORANGE)


# ─── PESTAÑA 3: KPIs Y SUPUESTOS ─────────────────────────────────────────────

def build_kpis(wb, data):
    ws = wb.create_sheet("KPIs y Supuestos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30

    row = 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value="KPIs, ESTIMADOS Y SUPUESTOS")
    c.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill = make_fill(DARK_BLUE)
    c.alignment = make_align("center")
    row += 2

    # KPIs por medio
    ws.merge_cells(f"A{row}:C{row}")
    set_cell(ws, row, 1, "MARCO DE MÉTRICAS POR MEDIO", bold=True, bg=MID_BLUE, fg=WHITE, size=11, align="center")
    row += 1

    headers_kpi = ["Medio", "KPI Primario (Comprometible)", "KPI Secundario (Estimado)"]
    for i, h in enumerate(headers_kpi):
        set_cell(ws, row, i+1, h, bold=True, bg=ACCENT_BLUE, fg=WHITE, size=10, align="center")
    row += 1

    filas = data.get("filas", [])
    medios_kpi = {}
    for f in filas:
        medio = f.get("medio", "")
        if medio not in medios_kpi:
            medios_kpi[medio] = (f.get("kpi_primario", ""), f.get("kpi_secundario", ""))

    alt = True
    for medio, (kpi1, kpi2) in medios_kpi.items():
        bg = LIGHT_BLUE if alt else VERY_LIGHT
        alt = not alt
        set_cell(ws, row, 1, medio, bold=True, bg=bg)
        set_cell(ws, row, 2, kpi1, bg=bg)
        set_cell(ws, row, 3, kpi2, bg=bg)
        row += 1

    row += 1

    # Supuestos generales
    supuestos = data.get("supuestos_generales", [])
    if supuestos:
        ws.merge_cells(f"A{row}:C{row}")
        set_cell(ws, row, 1, "SUPUESTOS GENERALES", bold=True, bg=MID_BLUE, fg=WHITE, size=11, align="center")
        row += 1
        for sup in supuestos:
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"A{row}:C{row}")
            set_cell(ws, row, 1, f"• {sup}", bg=VERY_LIGHT, size=9)
            row += 1
        row += 1


# ─── PESTAÑA 4: NOTAS COMERCIALES ────────────────────────────────────────────

def build_notas(wb, data):
    ws = wb.create_sheet("Notas Comerciales")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 70

    row = 1
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value="NOTAS COMERCIALES Y CONDICIONES")
    c.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill = make_fill(DARK_BLUE)
    c.alignment = make_align("center")
    row += 2

    notas = data.get("notas_comerciales", [])
    if notas:
        set_cell(ws, row, 1, "CONDICIONES Y RESTRICCIONES", bold=True, bg=MID_BLUE, fg=WHITE, size=11)
        row += 1
        for nota in notas:
            ws.row_dimensions[row].height = 24
            set_cell(ws, row, 1, f"▸  {nota}", bg=VERY_LIGHT, size=10)
            row += 1
        row += 1

    # Nota de pie estándar
    std_notas = [
        "Esta propuesta es de carácter comercial confidencial y está destinada exclusivamente al cliente indicado.",
        "Los precios están en MXN e incluyen IVA, salvo indicación contraria.",
        "La propuesta está sujeta a disponibilidad de inventario al momento de la confirmación.",
        "Los CPMs y estimados de audiencia son de referencia y pueden variar por plataforma y temporada.",
        "Se requiere confirmación por escrito para iniciar la reserva de espacios.",
    ]
    set_cell(ws, row, 1, "NOTAS GENERALES", bold=True, bg=LIGHT_GRAY, size=10)
    row += 1
    for nota in std_notas:
        ws.row_dimensions[row].height = 18
        set_cell(ws, row, 1, f"• {nota}", bg=WHITE, size=9)
        row += 1


# ─── BUILDER PRINCIPAL ───────────────────────────────────────────────────────

def build_excel(data: dict, output_path: str):
    wb = openpyxl.Workbook()

    # Eliminar la hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_resumen(wb, data)
    build_tactica(wb, data)
    build_kpis(wb, data)
    build_notas(wb, data)

    # Activar pestaña de resumen al abrir
    wb.active = wb["Resumen Ejecutivo"]

    wb.save(output_path)
    print(f"✅ Excel generado: {output_path}")
    return output_path


# ─── EJEMPLO DE DATOS ────────────────────────────────────────────────────────

EXAMPLE_DATA = {
    "cliente": "Ejemplo Corp",
    "marca": "Marca XYZ",
    "periodo": "Junio–Julio 2025",
    "objetivo": "Incrementar awareness y consideración de la marca entre adultos 25–44 NSE ABC+ en CDMX y Monterrey, para el lanzamiento del nuevo producto Premium.",
    "elaborado_por": "Área Comercial",
    "fecha": "Mayo 2025",
    "presupuesto_total": 5000000,
    "resumen_ejecutivo": "Propuesta táctica integrada de 8 semanas que combina TV Lineal para alcance masivo, CTV para segmentación precisa y refuerzo en streaming, y Digital para consideración y conversión. Total: $4,800,000 MXN con enfoque de funnel completo.",

    "filas": [
        {
            "medio": "TV Lineal",
            "propiedad": "Azteca Uno",
            "formato": 'Spot 20"',
            "descripcion": "Spot en bloque comercial en prime time (7PM–10PM). Máximo alcance en TV abierta.",
            "objetivo": "Awareness masivo nacional",
            "etapa_funnel": "Top",
            "audiencia": "A18–49, NSE ABC+, hogares con TV abierta",
            "entregable": "Spots al aire",
            "cantidad": 120,
            "unidad_compra": "Spots",
            "precio_unitario": 42000,
            "cpm_estimado": "$28 / GRP",
            "kpi_primario": "GRP / Alcance",
            "kpi_secundario": "Frecuencia media",
            "racional": "Azteca Uno es la señal de mayor alcance en TV abierta con cobertura nacional. Prime time garantiza exposición de 8–12M de hogares por noche.",
            "supuestos": "Precios de lista, sujeto a disponibilidad prime time. Requiere 10 días hábiles de anticipación."
        },
        {
            "medio": "TV Lineal",
            "propiedad": "Azteca 7",
            "formato": "Cortinilla con Mención",
            "descripcion": "Cortinilla de 10' con integración visual + mención verbal del conductor al inicio y cierre de bloque.",
            "objetivo": "Brand Association + Recall",
            "etapa_funnel": "Top",
            "audiencia": "A25–54, perfil mixto, afición deportiva y entretenimiento",
            "entregable": "Cortinillas emitidas",
            "cantidad": 80,
            "unidad_compra": "Unidades",
            "precio_unitario": 8500,
            "cpm_estimado": "—",
            "kpi_primario": "Brand Recall",
            "kpi_secundario": "Brand Association",
            "racional": "La cortinilla con mención en Azteca 7 combina presencia visual y endoso del conductor en un formato de alta atención y bajo costo.",
            "supuestos": "Sujeto a disponibilidad editorial. Creatividad requerida 5 días hábiles antes."
        },
        {
            "medio": "CTV",
            "propiedad": "Disney+ / Pluto TV",
            "formato": 'Pre-Roll Non-Skippable 30"',
            "descripcion": "Video de 30 segundos que se reproduce obligatoriamente antes del contenido. 95–99% de VCR garantizado.",
            "objetivo": "Awareness y message delivery en streaming",
            "etapa_funnel": "Top",
            "audiencia": "A25–45, NSE AB, usuarios activos de streaming en CDMX, GDL, MTY",
            "entregable": "Impresiones",
            "cantidad": 2000000,
            "unidad_compra": "Impresiones",
            "precio_unitario": 0.25,
            "cpm_estimado": "$250 CPM",
            "kpi_primario": "VCR / Reach incremental",
            "kpi_secundario": "Brand Awareness lift",
            "racional": "Disney+ y Pluto TV son las plataformas de mayor crecimiento en LATAM. El pre-roll non-skip en entorno premium garantiza 100% de visualización del mensaje.",
            "supuestos": "CPM es de lista. Puede negociarse a $200-220 con volumen. Mínimo 500K impresiones por campaña."
        },
        {
            "medio": "CTV",
            "propiedad": "Pluto TV / Tubi",
            "formato": "Channel Sponsorship FAST",
            "descripcion": "Patrocinio de canal temático completo. 'Este canal es presentado por [Marca]'. Presencia en todos los cortes del canal durante 4 semanas.",
            "objetivo": "Brand Association + Presencia sostenida en FAST",
            "etapa_funnel": "Top",
            "audiencia": "Usuarios de FAST TV, NSE B+C, 25–54 años",
            "entregable": "Semanas de patrocinio",
            "cantidad": 4,
            "unidad_compra": "Semanas",
            "precio_unitario": 85000,
            "cpm_estimado": "~$90 CPM",
            "kpi_primario": "Brand Association",
            "kpi_secundario": "Share of Voice en canal",
            "racional": "El patrocinio de canal en FAST ofrece presencia de marca de alta frecuencia a costo eficiente. Ideal para categorías de consumo masivo.",
            "supuestos": "Disponibilidad de canal sujeta a confirmación. Creatividades: logo de marca + cuña de 5'."
        },
        {
            "medio": "Digital",
            "propiedad": "Meta (Facebook + Instagram)",
            "formato": "Video In-Feed + Stories",
            "descripcion": "Video 15' en feed de noticias (1:1 y 9:16 para Stories). Optimizado para video views y alcance.",
            "objetivo": "Consideración y engagement con el mensaje de marca",
            "etapa_funnel": "Mid",
            "audiencia": "M/H 25–44, NSE ABC+, intereses: tecnología, lifestyle, consumo premium. CDMX, GDL, MTY.",
            "entregable": "Impresiones / Video Views",
            "cantidad": 5000000,
            "unidad_compra": "Impresiones",
            "precio_unitario": 0.12,
            "cpm_estimado": "$120 CPM",
            "kpi_primario": "Video Views / VTR",
            "kpi_secundario": "CTR / Post Engagement",
            "racional": "Meta permite segmentación hiper-precisa por intereses y comportamiento. El formato In-Feed Video captura atención en el momento de mayor intención de consumo de contenido.",
            "supuestos": "Creatividades en formato 1:1 y 9:16. A/B test de 2 versiones recomendado. Frequency cap: 3x por semana."
        },
        {
            "medio": "Digital",
            "propiedad": "YouTube",
            "formato": 'TrueView Skippable + Bumper 6"',
            "descripcion": "Pre-roll de 30' skippable a los 5'. Complementado con bumpers 6' de refuerzo y retargeting para quienes no saltaron.",
            "objetivo": "Consideración + Retargeting de audiencia interesada",
            "etapa_funnel": "Mid",
            "audiencia": "A18–44, usuarios de YouTube, segmentados por intereses en categoría",
            "entregable": "CPV / Vistas",
            "cantidad": 300000,
            "unidad_compra": "Vistas completadas (CPV)",
            "precio_unitario": 1.80,
            "cpm_estimado": "$80 CPM / $1.80 CPV",
            "kpi_primario": "CPV / VTR",
            "kpi_secundario": "Search Lift / Brand Recall",
            "racional": "YouTube TrueView filtra automáticamente la audiencia de mayor intención. Solo se paga cuando el usuario elige ver el anuncio completo. Los bumpers refuerzan el recall a muy bajo costo.",
            "supuestos": "Creatividad principal 30'. Bumper 6' de refuerzo. Brand Lift Study opcional (+$50K)."
        },
        {
            "medio": "Radio",
            "propiedad": "Grupo Fórmula / W Radio",
            "formato": 'Spot 30" + Mención en Vivo',
            "descripcion": "Spot de 30' en drive-time matutino (7AM–10AM) + 1 mención en vivo por el conductor del programa estelar por semana.",
            "objetivo": "Awareness en movilidad + Credibilidad por endoso del conductor",
            "etapa_funnel": "Top",
            "audiencia": "Adultos 30–55, NSE B+C, perfil de movilidad, escucha en auto",
            "entregable": "Spots + Menciones en vivo",
            "cantidad": 60,
            "unidad_compra": "Spots + Menciones",
            "precio_unitario": 12000,
            "cpm_estimado": "$45 CPM",
            "kpi_primario": "GRP / Cobertura",
            "kpi_secundario": "Brand Recall / Mención Orgánica",
            "racional": "El drive-time de Fórmula y W Radio concentra la audiencia de mayor NSE en movilidad. La mención en vivo del conductor potencia el mensaje con credibilidad editorial.",
            "supuestos": "Paquete combinado spots + menciones. Requiere guión de mención aprobado 3 días antes."
        },
    ],

    "supuestos_generales": [
        "Todos los precios en MXN con IVA incluido.",
        "Los CPMs son estimados de referencia y pueden variar ±15% por negociación y temporada.",
        "La propuesta está sujeta a disponibilidad de inventario al momento de la confirmación.",
        "Los GRPs de TV Lineal son estimados con base en histórico de ratings. Nielsen/Ibope es la fuente oficial.",
        "Los alcances en plataformas digitales son estimados según promedios de la plataforma para el target definido.",
        "Se requieren creatividades aprobadas con 5 días hábiles de anticipación al inicio de campaña.",
        "La propuesta no incluye producción de materiales publicitarios.",
    ],

    "notas_comerciales": [
        "Propuesta válida hasta el 15 de junio de 2025. Pasada esta fecha se requiere reconfirmación de precios.",
        "El patrocinio de canal en FAST (Pluto TV / Tubi) requiere mínimo 4 semanas de compromiso.",
        "Para CTV, el mínimo de compra es 500,000 impresiones por plataforma por campaña.",
        "El paquete de radio incluye 1 mención por semana en el programa estelar, sujeto a agenda editorial.",
        "Se puede negociar un bono de 10% en impresiones digitales con compromiso de 8 semanas.",
        "Contacto comercial: KAM Nombre Apellido — contacto@empresa.com — Tel: 55 0000 0000",
    ]
}


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Genera el Excel táctico comercial de medios")
    parser.add_argument("--data", help="Ruta al archivo JSON con los datos de la propuesta")
    parser.add_argument("--output", default="propuesta_comercial.xlsx", help="Ruta del archivo Excel a generar")
    parser.add_argument("--example", action="store_true", help="Genera un Excel de ejemplo")
    args = parser.parse_args()

    if args.example or not args.data:
        print("📊 Generando Excel de ejemplo...")
        build_excel(EXAMPLE_DATA, args.output)
    else:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
        build_excel(data, args.output)
